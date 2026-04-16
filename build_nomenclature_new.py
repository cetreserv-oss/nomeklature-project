from __future__ import annotations

import base64
import hashlib
import os
import re
import sys
import time
from copy import copy
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


REFERENCE_DIR = Path(r"E:\Projects\nomenklature project\reference")
SOURCE_KEYWORD = "\u041d\u043e\u043c\u0435\u043d\u043a\u043b\u0430\u0442\u0443\u0440\u0430"
REQUEST_KEYWORD = "\u0417\u0430\u043f\u0440\u043e\u0441 \u043f\u043e \u043c\u0430\u0440\u043a\u0430\u043c 1.4"
OUTPUT_NAME = "\u041d\u043e\u043c\u0435\u043d\u043a\u043b\u0430\u0442\u0443\u0440\u0430 New 2.xlsx"
LA_SOURCE_NAME = "\u041b\u0410"
SOURCE_COLUMN_COUNT = 20
REQUEST_TEMPLATE_SHEET_NAME = "\u041e\u0442\u043f\u0440\u0430\u0432\u043a\u0430 \u043c\u0430\u0440\u043e\u043a (\u043d\u043e\u043c\u0435\u043d\u043a\u043b\u0430\u0442\u0443\u0440\u0430)"
IMPORT_INFO_TEMPLATE_SHEET_NAME = "\u0421\u0432\u0435\u0434\u0435\u043d\u0438\u044f \u043e \u0432\u0432\u043e\u0437\u0435 (\u043d\u043e\u043c\u0435\u043d\u043a\u043b\u0430\u0442\u0443\u0440\u0430)"
WORKBOOK_WRITE_PASSWORD = "5623"
REGULAR_SHEET_PASSWORD = WORKBOOK_WRITE_PASSWORD
CORRECTION_SHEET_PASSWORD = "2356"
WRITE_RESERVATION_USER = "\u0430\u043b\u043a\u043e\u0433\u043e\u043b\u044c"

HEADER_CODE = "\u041a\u043e\u0434 \u0415\u0413\u0410\u0418\u0421"
HEADER_NAME = "\u041d\u043e\u043c\u0435\u043d\u043a\u043b\u0430\u0442\u0443\u0440\u0430"
CORRECTION_STATUS_HEADER = "\u0421\u0442\u0430\u0442\u0443\u0441"
FIX_STATUS_HEADER = "\u0421\u0442\u0430\u0442\u0443\u0441 \u0432 \u0415\u0413\u0410\u0418\u0421"
CORRECTION_STATUS_VALUES = [
    "\u041a\u043e\u0440\u0440\u0435\u043a\u0442\u0438\u0440\u043e\u0432\u043a\u0430 \u043f\u0440\u043e\u0432\u0435\u0434\u0435\u043d\u0430",
    "\u041e\u0436\u0438\u0434\u0430\u0435\u0442 \u043a\u043e\u0440\u0440\u0435\u043a\u0442\u0438\u0440\u043e\u0432\u043a\u0438",
    "\u041a\u043e\u0440\u0440\u0435\u043a\u0442\u0438\u0440\u043e\u0432\u043a\u0430 \u043d\u0435\u0430\u043a\u0442\u0443\u0430\u043b\u044c\u043d\u0430",
]
CORRECTION_COLUMN_COUNT = 10

SEND_DELETE_COLUMNS = [13, 11, 10, 9, 8]
FIX_DELETE_COLUMNS = [18]
SHIPPING_INVOICE_HEADER = "\u041d\u0430\u043a\u043b\u0430\u0434\u043d\u0430\u044f \u043e\u0442\u043f\u0440\u0430\u0432\u043a\u0438"
SHIPPING_INVOICE_HEADER_TYPO = "\u041d\u0430\u043a\u043b\u0430\u0434\u043d\u0430\u044f \u043e\u043f\u0440\u0430\u0432\u043a\u0438"

SHEET_SPECS = [
    {
        "source": "TK",
        "target": "TK \u043e\u0442\u043f\u0440\u0430\u0432\u043a\u0430 \u043c\u0430\u0440\u043a\u0438",
        "kind": "send",
        "table": "TkSend",
    },
    {
        "source": None,
        "target": "TK \u041a\u041e\u0420\u0420\u0415\u041a\u0422\u0418\u0420\u041e\u0412\u041a\u0410 \u043e\u0442\u043f\u0440\u0430\u0432\u043a\u0438 \u043c\u0430\u0440\u043a\u0438",
        "kind": "correction",
        "table": "TkSendCorrection",
    },
    {
        "source": "TK",
        "target": "TK \u0444\u0438\u043a\u0441\u0430\u0446\u0438\u044f \u0441\u0432\u0435\u0434\u0435\u043d\u0438\u0439 \u043e \u0432\u0432\u043e\u0437\u0435",
        "kind": "fix",
        "table": "TkImportFix",
    },
    {
        "source": None,
        "target": "TK \u041a\u041e\u0420\u0420. \u0441\u0432\u0435\u0434\u0435\u043d\u0438\u0439 \u043e \u0432\u0432\u043e\u0437\u0435",
        "kind": "correction",
        "template": "import_info",
        "table": "TkImportInfoCorrection",
    },
    {
        "source": LA_SOURCE_NAME,
        "target": "LA \u043e\u0442\u043f\u0440\u0430\u0432\u043a\u0430 \u043c\u0430\u0440\u043a\u0438",
        "kind": "send",
        "table": "LaSend",
    },
    {
        "source": None,
        "target": "LA \u041a\u041e\u0420\u0420\u0415\u041a\u0422\u0418\u0420\u041e\u0412\u041a\u0410 \u043e\u0442\u043f\u0440\u0430\u0432\u043a\u0438 \u043c\u0430\u0440\u043a\u0438",
        "kind": "correction",
        "table": "LaSendCorrection",
    },
    {
        "source": LA_SOURCE_NAME,
        "target": "LA \u0444\u0438\u043a\u0441\u0430\u0446\u0438\u044f \u0441\u0432\u0435\u0434\u0435\u043d\u0438\u0439 \u043e \u0432\u0432\u043e\u0437\u0435",
        "kind": "fix",
        "table": "LaImportFix",
    },
    {
        "source": None,
        "target": "LA \u041a\u041e\u0420\u0420. \u0441\u0432\u0435\u0434\u0435\u043d\u0438\u0439 \u043e \u0432\u0432\u043e\u0437\u0435",
        "kind": "correction",
        "template": "import_info",
        "table": "LaImportInfoCorrection",
    },
]


def find_source_workbook(reference_dir: Path) -> Path:
    candidates = sorted(
        path
        for path in reference_dir.glob("*.xlsx")
        if SOURCE_KEYWORD in path.stem
        and " New" not in path.stem
        and path.name != OUTPUT_NAME
        and not path.name.startswith("~$")
    )
    if not candidates:
        raise FileNotFoundError(
            f"\u041d\u0435 \u043d\u0430\u0439\u0434\u0435\u043d \u0438\u0441\u0445\u043e\u0434\u043d\u044b\u0439 \u0444\u0430\u0439\u043b \u0441 '{SOURCE_KEYWORD}' \u0432 {reference_dir}"
        )
    if len(candidates) > 1:
        names = ", ".join(path.name for path in candidates)
        raise RuntimeError(
            "\u041d\u0430\u0439\u0434\u0435\u043d\u043e \u043d\u0435\u0441\u043a\u043e\u043b\u044c\u043a\u043e "
            f"\u043f\u043e\u0434\u0445\u043e\u0434\u044f\u0449\u0438\u0445 \u0444\u0430\u0439\u043b\u043e\u0432, \u043d\u0443\u0436\u0435\u043d \u043e\u0434\u0438\u043d \u0438\u0441\u0442\u043e\u0447\u043d\u0438\u043a: {names}"
        )
    return candidates[0]


def find_request_workbook(reference_dir: Path) -> Path:
    candidates = sorted(
        path
        for path in reference_dir.glob("*.xlsm")
        if REQUEST_KEYWORD in path.stem and not path.name.startswith("~$")
    )
    if not candidates:
        raise FileNotFoundError(
            f"\u041d\u0435 \u043d\u0430\u0439\u0434\u0435\u043d \u0448\u0430\u0431\u043b\u043e\u043d \u0441 '{REQUEST_KEYWORD}' \u0432 {reference_dir}"
        )
    return candidates[-1]


def output_path() -> Path:
    return REFERENCE_DIR / OUTPUT_NAME


def load_correction_template(path: Path, sheet_name: str, data_column_count: int) -> dict:
    workbook = load_workbook(path, read_only=False, data_only=False, keep_vba=True)
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise RuntimeError(
            f"\u0412 \u0444\u0430\u0439\u043b\u0435 '{path.name}' \u043d\u0435 \u043d\u0430\u0439\u0434\u0435\u043d \u043b\u0438\u0441\u0442 '{sheet_name}'"
        )
    worksheet = workbook[sheet_name]
    headers = [
        worksheet.cell(row=1, column=column_index).value or ""
        for column_index in range(1, data_column_count + 1)
    ]
    widths = [
        worksheet.column_dimensions[worksheet.cell(row=1, column=column_index).column_letter].width
        or 12
        for column_index in range(1, data_column_count + 1)
    ]
    row_heights = [
        worksheet.row_dimensions[1].height or 30,
        worksheet.row_dimensions[2].height or 15,
    ]
    tab_color = None
    if worksheet.sheet_properties.tabColor is not None:
        tab_color = worksheet.sheet_properties.tabColor.rgb
    workbook.close()
    return {
        "headers": headers,
        "widths": widths,
        "row_heights": row_heights,
        "tab_color": tab_color,
    }


def wait_for_excel_ready(excel, timeout_seconds: int = 30) -> None:
    deadline = time.time() + timeout_seconds
    while time.time() < deadline:
        try:
            if excel.Ready:
                return
        except Exception:
            pass
        time.sleep(1)


def open_workbook_with_retry(excel, path: Path, **kwargs):
    last_error = None
    for _ in range(8):
        try:
            workbook = excel.Workbooks.Open(str(path), **kwargs)
            wait_for_excel_ready(excel)
            return workbook
        except Exception as exc:
            last_error = exc
            time.sleep(1)
    raise last_error


def close_workbook_safely(workbook, *, save_changes: bool) -> None:
    if workbook is None:
        return
    for _ in range(3):
        try:
            wait_for_excel_ready(workbook.Application)
            workbook.Close(SaveChanges=save_changes)
            return
        except Exception:
            time.sleep(1)


def worksheet(workbook, key):
    return workbook.Worksheets(key)


def last_filled_row(worksheet) -> int:
    xl_formulas = -4123
    xl_by_rows = 1
    xl_previous = 2
    last_cell = worksheet.Columns(1).Find(
        What="*",
        LookIn=xl_formulas,
        SearchOrder=xl_by_rows,
        SearchDirection=xl_previous,
    )
    return last_cell.Row if last_cell is not None else 1


def copy_source_range(source_sheet, target_sheet, row_count: int) -> None:
    source_sheet.Range(
        source_sheet.Cells(1, 1),
        source_sheet.Cells(row_count, SOURCE_COLUMN_COUNT),
    ).Copy(target_sheet.Cells(1, 1))

    for column_index in range(1, SOURCE_COLUMN_COUNT + 1):
        source_column = source_sheet.Columns(column_index)
        target_column = target_sheet.Columns(column_index)
        target_column.ColumnWidth = source_column.ColumnWidth
        target_column.Hidden = source_column.Hidden

    for row_index in range(1, row_count + 1):
        source_row = source_sheet.Rows(row_index)
        target_row = target_sheet.Rows(row_index)
        target_row.RowHeight = source_row.RowHeight
        target_row.Hidden = source_row.Hidden

    target_sheet.Tab.Color = source_sheet.Tab.Color


def normalize_code_columns(worksheet, source_name: str) -> None:
    if source_name == "TK":
        row_count = last_filled_row(worksheet)
        if row_count >= 2:
            code_values = worksheet.Range(
                worksheet.Cells(2, 3),
                worksheet.Cells(row_count, 3),
            ).Value
            nomenclature_values = worksheet.Range(
                worksheet.Cells(2, 2),
                worksheet.Cells(row_count, 2),
            ).Value
            worksheet.Range(
                worksheet.Cells(2, 2),
                worksheet.Cells(row_count, 2),
            ).Value = code_values
            worksheet.Range(
                worksheet.Cells(2, 3),
                worksheet.Cells(row_count, 3),
            ).Value = nomenclature_values

        width_b = worksheet.Columns(2).ColumnWidth
        width_c = worksheet.Columns(3).ColumnWidth
        worksheet.Columns(2).ColumnWidth = width_c
        worksheet.Columns(3).ColumnWidth = width_b

    worksheet.Cells(1, 2).Value = HEADER_CODE
    worksheet.Cells(1, 3).Value = HEADER_NAME


def trim_columns(worksheet, kind: str) -> int:
    delete_columns = SEND_DELETE_COLUMNS if kind == "send" else FIX_DELETE_COLUMNS
    for column_index in delete_columns:
        worksheet.Columns(column_index).Delete()
    return SOURCE_COLUMN_COUNT - len(delete_columns)


def normalize_headers(worksheet, column_count: int) -> None:
    for column_index in range(1, column_count + 1):
        value = str(worksheet.Cells(1, column_index).Value or "").strip()
        if value == SHIPPING_INVOICE_HEADER_TYPO:
            worksheet.Cells(1, column_index).Value = SHIPPING_INVOICE_HEADER


def find_header_column(worksheet, header: str, column_count: int):
    for column_index in range(1, column_count + 1):
        value = str(worksheet.Cells(1, column_index).Value or "").strip()
        if value.lower() == header.lower():
            return column_index
    return None


def apply_filter_and_freeze(worksheet, row_count: int, column_count: int) -> None:
    worksheet.Activate()
    active_window = worksheet.Application.ActiveWindow
    active_window.SplitColumn = 0
    active_window.SplitRow = 1
    active_window.FreezePanes = False
    active_window.FreezePanes = True


def rgb_color(red: int, green: int, blue: int) -> int:
    return red + (green << 8) + (blue << 16)


def ensure_table(worksheet, row_count: int, column_count: int, table_name: str) -> None:
    while worksheet.ListObjects.Count > 0:
        worksheet.ListObjects(1).Unlist()

    table = worksheet.ListObjects.Add(
        1,
        worksheet.Range(
            worksheet.Cells(1, 1),
            worksheet.Cells(row_count, column_count),
        ),
        None,
        1,
    )
    table.Name = table_name
    table.TableStyle = ""


def prepare_correction_status_dictionary(workbook) -> None:
    helper_sheet = worksheet(workbook, 1)
    helper_sheet.Cells(1, 3).Value = CORRECTION_STATUS_HEADER
    for index, value in enumerate(CORRECTION_STATUS_VALUES, start=2):
        helper_sheet.Cells(index, 3).Value = value
    helper_sheet.Columns(3).ColumnWidth = 32


def add_status_validation(worksheet, validation_formula: str) -> None:
    status_range = worksheet.Range(
        worksheet.Cells(2, CORRECTION_COLUMN_COUNT),
        worksheet.Cells(worksheet.Rows.Count, CORRECTION_COLUMN_COUNT),
    )
    status_range.Validation.Delete()
    status_range.Validation.Add(
        Type=3,
        AlertStyle=1,
        Operator=1,
        Formula1=validation_formula,
    )
    status_range.Validation.IgnoreBlank = True
    status_range.Validation.InCellDropdown = True


def add_status_formatting(worksheet) -> None:
    xl_cell_value = 1
    xl_equal = 3
    status_range = worksheet.Range(
        worksheet.Cells(2, CORRECTION_COLUMN_COUNT),
        worksheet.Cells(worksheet.Rows.Count, CORRECTION_COLUMN_COUNT),
    )
    status_range.FormatConditions.Delete()

    green_rule = status_range.FormatConditions.Add(
        Type=xl_cell_value,
        Operator=xl_equal,
        Formula1='="{}"'.format(CORRECTION_STATUS_VALUES[0]),
    )
    green_rule.Interior.Color = rgb_color(198, 239, 206)
    green_rule.Font.Color = rgb_color(0, 97, 0)

    yellow_rule = status_range.FormatConditions.Add(
        Type=xl_cell_value,
        Operator=xl_equal,
        Formula1='="{}"'.format(CORRECTION_STATUS_VALUES[1]),
    )
    yellow_rule.Interior.Color = rgb_color(255, 235, 156)
    yellow_rule.Font.Color = rgb_color(156, 101, 0)

    red_rule = status_range.FormatConditions.Add(
        Type=xl_cell_value,
        Operator=xl_equal,
        Formula1='="{}"'.format(CORRECTION_STATUS_VALUES[2]),
    )
    red_rule.Interior.Color = rgb_color(255, 199, 206)
    red_rule.Font.Color = rgb_color(156, 0, 6)


def build_correction_sheet(worksheet, template: dict, validation_formula: str, table_name: str) -> None:
    for column_index, header in enumerate(template["headers"], start=1):
        worksheet.Cells(1, column_index).Value = header

    worksheet.Cells(1, 10).Value = CORRECTION_STATUS_HEADER
    worksheet.Cells(2, 10).FormulaR1C1 = (
        '=IF(COUNTA(RC1:RC9)>0,"{}","")'.format(CORRECTION_STATUS_VALUES[1])
    )

    header_range = worksheet.Range(worksheet.Cells(1, 1), worksheet.Cells(1, 10))
    header_range.Interior.Color = 65535
    header_range.Font.Bold = True
    header_range.HorizontalAlignment = -4108
    header_range.VerticalAlignment = -4108
    header_range.WrapText = True

    for column_index, width in enumerate(template["widths"], start=1):
        worksheet.Columns(column_index).ColumnWidth = width
    worksheet.Columns(10).ColumnWidth = 26
    worksheet.Rows(1).RowHeight = template["row_heights"][0]
    worksheet.Rows(2).RowHeight = template["row_heights"][1]
    if template["tab_color"]:
        rgb = template["tab_color"][-6:]
        worksheet.Tab.Color = rgb_color(
            int(rgb[0:2], 16),
            int(rgb[2:4], 16),
            int(rgb[4:6], 16),
        )

    ensure_table(worksheet, 2, CORRECTION_COLUMN_COUNT, table_name)
    add_status_validation(worksheet, validation_formula)
    add_status_formatting(worksheet)


def protect_sheet(
    worksheet,
    row_count: int,
    column_count: int,
    *,
    password: str,
    allow_delete_rows: bool,
    allow_insert_rows: bool,
    unlock_data_cells: bool,
    unlock_status_cells: bool,
    unlocked_headers: tuple[str, ...] = (),
) -> None:
    worksheet.Cells.Locked = True
    if unlock_status_cells:
        editable_last_row = max(row_count, 2)
        worksheet.Range(
            worksheet.Cells(2, CORRECTION_COLUMN_COUNT),
            worksheet.Cells(editable_last_row, CORRECTION_COLUMN_COUNT),
        ).Locked = False
    elif unlock_data_cells:
        worksheet.Range(
            worksheet.Cells(2, 1),
            worksheet.Cells(worksheet.Rows.Count, worksheet.Columns.Count),
        ).Locked = False
    for header in unlocked_headers:
        column_index = find_header_column(worksheet, header, column_count)
        if column_index is None:
            raise RuntimeError(f"\u041d\u0435 \u043d\u0430\u0439\u0434\u0435\u043d \u0441\u0442\u043e\u043b\u0431\u0435\u0446 '{header}' \u043d\u0430 \u043b\u0438\u0441\u0442\u0435 '{worksheet.Name}'")
        worksheet.Range(
            worksheet.Cells(2, column_index),
            worksheet.Cells(worksheet.Rows.Count, column_index),
        ).Locked = False

    worksheet.EnableSelection = 0
    worksheet.Protect(
        Password=password,
        DrawingObjects=True,
        Contents=True,
        Scenarios=True,
        UserInterfaceOnly=False,
        AllowFormattingCells=False,
        AllowFormattingColumns=True,
        AllowFormattingRows=False,
        AllowInsertingColumns=False,
        AllowInsertingRows=allow_insert_rows,
        AllowInsertingHyperlinks=False,
        AllowDeletingColumns=False,
        AllowDeletingRows=allow_delete_rows,
        AllowSorting=False,
        AllowFiltering=True,
        AllowUsingPivotTables=False,
    )


def openpyxl_last_filled_row(worksheet) -> int:
    for row_index in range(worksheet.max_row, 0, -1):
        if worksheet.cell(row=row_index, column=1).value not in (None, ""):
            return row_index
    return 1


def copy_openpyxl_cell(source_cell, target_cell) -> None:
    target_cell.value = source_cell.value
    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format
    if source_cell.hyperlink:
        target_cell._hyperlink = copy(source_cell.hyperlink)
    if source_cell.comment:
        target_cell.comment = copy(source_cell.comment)


def copy_openpyxl_source_range(source_sheet, target_sheet, row_count: int) -> None:
    for row in range(1, row_count + 1):
        for column in range(1, SOURCE_COLUMN_COUNT + 1):
            copy_openpyxl_cell(
                source_sheet.cell(row=row, column=column),
                target_sheet.cell(row=row, column=column),
            )

    for column in range(1, SOURCE_COLUMN_COUNT + 1):
        letter = get_column_letter(column)
        source_dimension = source_sheet.column_dimensions[letter]
        target_dimension = target_sheet.column_dimensions[letter]
        target_dimension.width = source_dimension.width
        target_dimension.hidden = source_dimension.hidden

    for row in range(1, row_count + 1):
        source_dimension = source_sheet.row_dimensions[row]
        target_dimension = target_sheet.row_dimensions[row]
        target_dimension.height = source_dimension.height
        target_dimension.hidden = source_dimension.hidden

    if source_sheet.sheet_properties.tabColor is not None:
        target_sheet.sheet_properties.tabColor = source_sheet.sheet_properties.tabColor.rgb


def normalize_openpyxl_code_columns(worksheet, source_name: str) -> None:
    if source_name == "TK":
        row_count = openpyxl_last_filled_row(worksheet)
        for row in range(2, row_count + 1):
            column_b = worksheet.cell(row=row, column=2)
            column_c = worksheet.cell(row=row, column=3)
            column_b.value, column_c.value = column_c.value, column_b.value

        width_b = worksheet.column_dimensions["B"].width
        width_c = worksheet.column_dimensions["C"].width
        worksheet.column_dimensions["B"].width = width_c
        worksheet.column_dimensions["C"].width = width_b

    worksheet.cell(row=1, column=2).value = HEADER_CODE
    worksheet.cell(row=1, column=3).value = HEADER_NAME


def trim_openpyxl_columns(worksheet, kind: str) -> int:
    delete_columns = SEND_DELETE_COLUMNS if kind == "send" else FIX_DELETE_COLUMNS
    for column_index in sorted(delete_columns, reverse=True):
        worksheet.delete_cols(column_index)
    return SOURCE_COLUMN_COUNT - len(delete_columns)


def normalize_openpyxl_headers(worksheet, column_count: int) -> None:
    for column_index in range(1, column_count + 1):
        value = str(worksheet.cell(row=1, column=column_index).value or "").strip()
        if value == SHIPPING_INVOICE_HEADER_TYPO:
            worksheet.cell(row=1, column=column_index).value = SHIPPING_INVOICE_HEADER


def find_openpyxl_header_column(worksheet, header: str, column_count: int) -> int | None:
    for column_index in range(1, column_count + 1):
        value = str(worksheet.cell(row=1, column=column_index).value or "").strip()
        if value.lower() == header.lower():
            return column_index
    return None


def ensure_openpyxl_table(worksheet, row_count: int, column_count: int, table_name: str) -> None:
    # Excel COM rejects some openpyxl-created ListObjects in this workbook.
    # Keep a plain protected range with AutoFilter; VBA falls back to row append.
    worksheet.tables.clear()


def prepare_openpyxl_status_dictionary(workbook) -> None:
    helper_sheet = workbook.worksheets[0]
    helper_sheet.cell(row=1, column=3).value = CORRECTION_STATUS_HEADER
    for index, value in enumerate(CORRECTION_STATUS_VALUES, start=2):
        helper_sheet.cell(row=index, column=3).value = value
    helper_sheet.column_dimensions["C"].width = 32


def list_formula(sheet_name: str, column_letter: str, first_row: int, last_row: int) -> str:
    escaped_name = sheet_name.replace("'", "''")
    return f"'{escaped_name}'!${column_letter}${first_row}:${column_letter}${last_row}"


def add_openpyxl_list_validation(worksheet, formula1: str, range_reference: str) -> None:
    validation = DataValidation(
        type="list",
        formula1=formula1,
        allow_blank=True,
        showErrorMessage=True,
        showInputMessage=True,
    )
    validation.add(range_reference)
    worksheet.add_data_validation(validation)


def add_openpyxl_status_formatting(worksheet, range_reference: str) -> None:
    colors = [
        ("C6EFCE", "006100"),
        ("FFEB9C", "9C6500"),
        ("FFC7CE", "9C0006"),
    ]
    for value, (fill_color, font_color) in zip(CORRECTION_STATUS_VALUES, colors):
        worksheet.conditional_formatting.add(
            range_reference,
            CellIsRule(
                operator="equal",
                formula=[f'"{value}"'],
                fill=PatternFill(fill_type="solid", fgColor=fill_color),
                font=Font(color=font_color),
            ),
        )


def build_openpyxl_correction_sheet(
    worksheet,
    template: dict,
    validation_formula: str,
    table_name: str,
) -> None:
    data_column_count = len(template["headers"])
    status_column = data_column_count + 1
    for column_index, header in enumerate(template["headers"], start=1):
        worksheet.cell(row=1, column=column_index).value = header

    worksheet.cell(row=1, column=status_column).value = CORRECTION_STATUS_HEADER
    worksheet.cell(row=2, column=status_column).value = (
        f'=IF(COUNTA(A2:{get_column_letter(data_column_count)}2)>0,"{CORRECTION_STATUS_VALUES[1]}","")'
    )

    header_fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    for column in range(1, status_column + 1):
        cell = worksheet.cell(row=1, column=column)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for column_index, width in enumerate(template["widths"], start=1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = width
    worksheet.column_dimensions[get_column_letter(status_column)].width = 26
    worksheet.row_dimensions[1].height = template["row_heights"][0]
    worksheet.row_dimensions[2].height = template["row_heights"][1]
    if template["tab_color"]:
        worksheet.sheet_properties.tabColor = template["tab_color"]

    ensure_openpyxl_table(worksheet, 2, status_column, table_name)
    status_range = f"{get_column_letter(status_column)}2:{get_column_letter(status_column)}1048576"
    add_openpyxl_list_validation(worksheet, validation_formula, status_range)
    add_openpyxl_status_formatting(worksheet, status_range)


def apply_openpyxl_filter_and_freeze(worksheet, row_count: int, column_count: int) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(column_count)}{max(row_count, 1)}"


def protect_openpyxl_sheet(
    worksheet,
    row_count: int,
    column_count: int,
    *,
    password: str,
    allow_delete_rows: bool,
    allow_insert_rows: bool,
    unlock_data_cells: bool,
    unlocked_headers: tuple[str, ...] = (),
) -> None:
    locked = Protection(locked=True)
    unlocked = Protection(locked=False)

    for row in worksheet.iter_rows(
        min_row=1,
        max_row=max(row_count, 1),
        min_col=1,
        max_col=column_count,
    ):
        for cell in row:
            cell.protection = locked

    if unlock_data_cells:
        for column_index in range(1, column_count + 1):
            worksheet.column_dimensions[get_column_letter(column_index)].protection = unlocked
        for row in worksheet.iter_rows(
            min_row=2,
            max_row=max(row_count, 2),
            min_col=1,
            max_col=column_count,
        ):
            for cell in row:
                cell.protection = unlocked

    for header in unlocked_headers:
        column_index = find_openpyxl_header_column(worksheet, header, column_count)
        if column_index is None:
            raise RuntimeError(
                f"\u041d\u0435 \u043d\u0430\u0439\u0434\u0435\u043d \u0441\u0442\u043e\u043b\u0431\u0435\u0446 '{header}' \u043d\u0430 \u043b\u0438\u0441\u0442\u0435 '{worksheet.title}'"
            )
        worksheet.column_dimensions[get_column_letter(column_index)].protection = unlocked
        for row_index in range(2, max(row_count, 2) + 1):
            worksheet.cell(row=row_index, column=column_index).protection = unlocked
        worksheet.cell(row=1, column=column_index).protection = locked

    worksheet.protection.sheet = True
    worksheet.protection.set_password(password)
    worksheet.protection.formatCells = True
    worksheet.protection.formatColumns = False
    worksheet.protection.formatRows = True
    worksheet.protection.insertColumns = True
    worksheet.protection.insertRows = not allow_insert_rows
    worksheet.protection.insertHyperlinks = True
    worksheet.protection.deleteColumns = True
    worksheet.protection.deleteRows = not allow_delete_rows
    worksheet.protection.sort = True
    worksheet.protection.autoFilter = False
    worksheet.protection.pivotTables = True
    worksheet.protection.objects = True
    worksheet.protection.scenarios = True


def excel_sha512_hash(password: str, salt: bytes, spin_count: int = 100000) -> str:
    digest = hashlib.sha512(salt + password.encode("utf-16le")).digest()
    for index in range(spin_count):
        digest = hashlib.sha512(digest + index.to_bytes(4, "little")).digest()
    return base64.b64encode(digest).decode("ascii")


def set_write_reservation_password(path: Path, password: str) -> None:
    temp_path = path.with_suffix(".tmp.xlsx")
    salt = os.urandom(16)
    spin_count = 100000
    file_sharing = (
        f'<fileSharing userName="{WRITE_RESERVATION_USER}" algorithmName="SHA-512" '
        f'hashValue="{excel_sha512_hash(password, salt, spin_count)}" '
        f'saltValue="{base64.b64encode(salt).decode("ascii")}" '
        f'spinCount="{spin_count}"/>'
    )

    with ZipFile(path, "r") as source, ZipFile(temp_path, "w", ZIP_DEFLATED) as target:
        for item in source.infolist():
            data = source.read(item.filename)
            if item.filename == "xl/workbook.xml":
                xml = data.decode("utf-8")
                xml = re.sub(r"<fileSharing\b[^>]*/>", "", xml)
                if re.search(r"<workbookPr\b[^>]*/>", xml):
                    xml = re.sub(
                        r"(<workbookPr\b[^>]*/>)",
                        file_sharing + r"\1",
                        xml,
                        count=1,
                    )
                elif re.search(r"<fileVersion\b[^>]*/>", xml):
                    xml = re.sub(
                        r"(<fileVersion\b[^>]*/>)",
                        r"\1" + file_sharing,
                        xml,
                        count=1,
                    )
                else:
                    xml = re.sub(r"(<workbook\b[^>]*>)", r"\1" + file_sharing, xml, count=1)
                data = xml.encode("utf-8")
            target.writestr(item, data)

    temp_path.replace(path)


def add_fix_sheet_validations(workbook, worksheet, row_count: int, column_count: int) -> None:
    helper_sheet = workbook.worksheets[0]
    helper_name = helper_sheet.title
    status_column = find_openpyxl_header_column(worksheet, FIX_STATUS_HEADER, column_count)
    import_column = find_openpyxl_header_column(
        worksheet,
        "\u0417\u0430\u0444\u0438\u043a\u0441\u0438\u0440\u043e\u0432\u0430\u0442\u044c \u0432\u0432\u043e\u0437",
        column_count,
    )

    if status_column is not None:
        letter = get_column_letter(status_column)
        add_openpyxl_list_validation(
            worksheet,
            list_formula(helper_name, "A", 2, 4),
            f"{letter}2:{letter}{max(row_count, 2)}",
        )
    if import_column is not None:
        letter = get_column_letter(import_column)
        add_openpyxl_list_validation(
            worksheet,
            list_formula(helper_name, "B", 2, 3),
            f"{letter}2:{letter}{max(row_count, 2)}",
        )


def build_workbook() -> Path:
    source_path = find_source_workbook(REFERENCE_DIR)
    request_path = find_request_workbook(REFERENCE_DIR)
    destination_path = output_path()

    if destination_path.exists():
        destination_path.unlink()
    correction_templates = {
        "send": load_correction_template(request_path, REQUEST_TEMPLATE_SHEET_NAME, 9),
        "import_info": load_correction_template(request_path, IMPORT_INFO_TEMPLATE_SHEET_NAME, 14),
    }

    workbook = load_workbook(source_path)
    prepare_openpyxl_status_dictionary(workbook)
    correction_validation_formula = '"' + ",".join(CORRECTION_STATUS_VALUES) + '"'

    for spec in SHEET_SPECS:
        new_sheet = workbook.create_sheet(spec["target"])

        if spec["kind"] == "correction":
            correction_template = correction_templates[spec.get("template", "send")]
            build_openpyxl_correction_sheet(
                new_sheet,
                correction_template,
                correction_validation_formula,
                spec["table"],
            )
            row_count = 2
            column_count = len(correction_template["headers"]) + 1
            password = CORRECTION_SHEET_PASSWORD
            allow_delete_rows = False
            allow_insert_rows = False
            unlock_data_cells = False
            unlocked_headers = (CORRECTION_STATUS_HEADER,)
        else:
            source_sheet = workbook[spec["source"]]
            row_count = openpyxl_last_filled_row(source_sheet)
            copy_openpyxl_source_range(source_sheet, new_sheet, row_count)
            normalize_openpyxl_code_columns(new_sheet, spec["source"])
            column_count = trim_openpyxl_columns(new_sheet, spec["kind"])
            normalize_openpyxl_headers(new_sheet, column_count)
            row_count = openpyxl_last_filled_row(new_sheet)
            ensure_openpyxl_table(new_sheet, row_count, column_count, spec["table"])

            if spec["kind"] == "fix":
                add_fix_sheet_validations(workbook, new_sheet, row_count, column_count)
                password = REGULAR_SHEET_PASSWORD
                allow_delete_rows = True
                allow_insert_rows = True
                unlock_data_cells = True
                unlocked_headers = ()
            else:
                password = REGULAR_SHEET_PASSWORD
                allow_delete_rows = True
                allow_insert_rows = True
                unlock_data_cells = True
                unlocked_headers = ()

        apply_openpyxl_filter_and_freeze(new_sheet, row_count, column_count)
        protect_openpyxl_sheet(
            new_sheet,
            row_count,
            column_count,
            password=password,
            allow_delete_rows=allow_delete_rows,
            allow_insert_rows=allow_insert_rows,
            unlock_data_cells=unlock_data_cells,
            unlocked_headers=unlocked_headers,
        )

    for source_sheet_name in ("TK", LA_SOURCE_NAME):
        if source_sheet_name in workbook.sheetnames:
            del workbook[source_sheet_name]

    workbook.properties.creator = WRITE_RESERVATION_USER
    workbook.properties.lastModifiedBy = WRITE_RESERVATION_USER
    workbook.save(destination_path)
    workbook.close()
    set_write_reservation_password(destination_path, WORKBOOK_WRITE_PASSWORD)

    return destination_path


def main() -> int:
    try:
        created_path = build_workbook()
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    print(created_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
